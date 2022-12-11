from datetime import timedelta
from pathlib import Path

from evsauto.spreadsheet import load_workbook, XLFormatter
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.cell import get_column_letter
import pandas as pd
import pyperclip


# suppress warnings
import warnings
warnings.simplefilter("ignore")


def debug_func(source_path, date):
    source_df = pd.read_excel(source_path, engine="openpyxl")
    source_df = source_df.convert_dtypes()

    source_df.drop(source_df.loc[
        source_df["1 - WO Owner"] != "DEVERDAN"
    ].index, inplace=True)
    # source_df.drop(
    #     source_df.loc[source_df["Sched. Start Date"] > date].index, inplace=True)

    print(source_df["Work Order"].count())


def generate_work_order_summary(source_path, target_path, sheet_to_columns, username, mc_date, date):
    source_df = pd.read_excel(source_path, engine="openpyxl")
    source_df = source_df.convert_dtypes()

    all_work_orders_df = source_df[
        ((source_df["1 - WO Owner"].eq(username)) |
         source_df["Assigned to on Activity (Top 8 activities)"].str.contains(username)) &
        (pd.to_datetime(source_df["Sched. Start Date"])
         <= pd.Timestamp(date.year, date.month, date.day))
    ]
    all_work_orders_df = all_work_orders_df.sort_values(
        by=["Sched. Start Date"], ascending=False)
    all_work_orders_df = all_work_orders_df[sheet_to_columns["Work Orders Involving Me"]]

    mc_type_filter = [
        "PdM - Work using a Predictive Tool",
        "PM - Preventative Maintenance"
    ]
    max_compliance_df = source_df[
        source_df.Type.isin(mc_type_filter) &
        (pd.to_datetime(source_df["PM Compliance Max"]).eq(
            pd.Timestamp(mc_date.year, mc_date.month, mc_date.day)))
    ]
    max_compliance_df = max_compliance_df[sheet_to_columns["Due by Midnight"]]

    backlog_df = source_df[
        (source_df.Type != "Training - Time for Training Activities (Sys Sched)") &
        (pd.to_datetime(source_df["Sched. Start Date"])
         < pd.Timestamp(date.year, date.month, date.day))
    ]
    backlog_df = backlog_df.drop(max_compliance_df.index)
    backlog_df = backlog_df[sheet_to_columns["Backlog"]]

    todays_work_df = source_df[
        (source_df.Type != "Training - Time for Training Activities (Sys Sched)") &
        (pd.to_datetime(source_df["Sched. Start Date"]).eq(
            pd.Timestamp(date.year, date.month, date.day)))
    ]
    todays_work_df = todays_work_df[sheet_to_columns["Scheduled for Shift"]]
    todays_work_df = todays_work_df.sort_values(
        by=["1 - WO Owner"], ascending=True)

    follow_ups_df = source_df[
        (source_df["Type"] ==
         "Follow-Up - Comes from a scheduled WO Checklist (PM, PdM)")
    ]
    follow_ups_df = follow_ups_df[sheet_to_columns["Follow Ups"]]

    training_df = source_df[
        (source_df["Type"] ==
         "Training - Time for Training Activities (Sys Sched)")
    ]
    training_df = training_df[sheet_to_columns["Training"]]

    if Path(target_path).exists():
        Path(target_path).unlink()  # delete

    with pd.ExcelWriter(target_path,
                        engine="xlsxwriter",
                        engine_kwargs={'options': {'strings_to_numbers': True}}) as writer:
        if all_work_orders_df["Work Order"].count() != 0:
            all_work_orders_df.to_excel(
                writer,
                sheet_name="Work Orders Involving Me",
                index=False
            )
        if todays_work_df["Work Order"].count() != 0:
            todays_work_df.to_excel(
                writer,
                sheet_name="Scheduled for Shift",
                index=False
            )
        if max_compliance_df["Work Order"].count() != 0:
            max_compliance_df.to_excel(
                writer,
                sheet_name="Due by Midnight",
                index=False
            )
        if backlog_df["Work Order"].count() != 0:
            backlog_df.to_excel(
                writer,
                sheet_name="Backlog",
                index=False
            )
        if follow_ups_df["Work Order"].count() != 0:
            follow_ups_df.to_excel(
                writer,
                sheet_name="Follow Ups",
                index=False
            )
        if training_df["Work Order"].count() != 0:
            training_df.to_excel(
                writer,
                sheet_name="Training",
                index=False
            )


def format_work_order_summary(report_path, sheet_to_columns, username):
    # use openpyxl to format the excel file
    with load_workbook(report_path) as wo_summary_report:

        # formatter init
        xlformat = XLFormatter(workbook=wo_summary_report)

        # loop though the workbook sheets
        for sheetname in wo_summary_report.sheetnames:
            sheet = wo_summary_report[sheetname]
            columns = sheet_to_columns[sheetname]

            centered_cols = [i for i, col in enumerate(columns) if col in [
                'Work Order', 'Priority Icon', 'Sched. Start Date', 'PM Compliance Min', 'PM Compliance Max', 'Reported By', 'Parent Work Order', 'Status', 'Department', 'Organization']]
            left_indent_cols = [i for i, col in enumerate(columns) if col in [
                'Description', 'Type', 'Equipment', 'Equipment Description', '1 - WO Owner', 'Assigned to on Activity (Top 8 activities)']]
            date_cols = [i for i, col in enumerate(columns) if col in [
                'Sched. Start Date', 'PM Compliance Min', 'PM Compliance Max']]

            # align columns
            xlformat.set_sheet(sheet).align_all(left_indent_cols)
            xlformat.align_all(centered_cols, align_settings={
                "horizontal": "center",
                "vertical": "center",
                "wrap_text": True
            })

            # autofit the column width
            xlformat.autofit_columns()

            # set the number format of the date columns
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for i in date_cols:
                    row[i].number_format = "m/d/yyyy"

            # set the header row height
            sheet.row_dimensions[1].height = 28

            # set the background color and font size of the header
            for cell in sheet[1]:
                cell.font = Font(
                    bold=True,
                    color='FFFFFFFF'
                )
                cell.fill = PatternFill(
                    fill_type='solid',
                    fgColor='FF244062'
                )

            # set the borders of the table
            thin_border = Side(border_style='thin', color='FF000000')

            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                for cell in row:
                    cell.border = Border(
                        left=thin_border,
                        right=thin_border,
                        top=thin_border,
                        bottom=thin_border
                    )

            # disable grid lines
            sheet.sheet_view.showGridLines = False

            # copy my work for the shift to clipboard
            if sheetname == "Scheduled for Shift":
                my_work = ""
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    if username in str(row[3].value):
                        for i in [0, 2]:
                            my_work += (str(row[i].value) + "\t")
                        my_work += "\n"
                pyperclip.copy(my_work)

            # freeze the first row
            sheet.freeze_panes = "A2"

            # setup a filter
            sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
