from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime, timedelta
import pandas as pd
import pyperclip
import os


import warnings
warnings.simplefilter("ignore")


def generate_work_order_summary(source_path, target_path, sheet_to_columns, date=datetime.today()):
    source_df = pd.read_excel(source_path, engine="openpyxl")
    source_df = source_df.convert_dtypes()

    mc_type_filter = [
        "PdM - Work using a Predictive Tool",
        "PM - Preventative Maintenance"
    ]
    mc_date_filter = date.strftime("%#m/%#d/%Y")
    tw_date_filter = date + timedelta(days=1)
    tw_date_filter = tw_date_filter.strftime("%#m/%#d/%Y")

    mc_df = source_df[
        source_df.Type.isin(mc_type_filter) &
        (source_df["PM Compliance Max"] == mc_date_filter)
    ]
    mc_df = mc_df[sheet_to_columns["Due by Midnight"]]

    blog_df = source_df[
        (source_df.Type != "Training - Time for Training Activities (Sys Sched)") &
        (source_df["Sched. Start Date"] <= date)
    ]
    blog_df = blog_df.drop(mc_df.index)
    blog_df = blog_df[sheet_to_columns["Backlog"]]

    tw_df = source_df[
        (source_df["Sched. Start Date"] == tw_date_filter) &
        (source_df.Type != "Training - Time for Training Activities (Sys Sched)")
    ]
    tw_df = tw_df[sheet_to_columns["Scheduled for Shift"]]
    tw_df = tw_df.sort_values(by=["1 - WO Owner"], ascending=True)

    fu_df = source_df[
        (source_df["Type"] ==
         "Follow-Up - Comes from a scheduled WO Checklist (PM, PdM)")
    ]
    fu_df = fu_df[sheet_to_columns["Follow Ups"]]

    tr_df = source_df[
        (source_df["Type"] ==
         "Training - Time for Training Activities (Sys Sched)")
    ]
    tr_df = tr_df[sheet_to_columns["Training"]]

    with pd.ExcelWriter(target_path,
                        engine="xlsxwriter",
                        engine_kwargs={'options': {'strings_to_numbers': True}}) as writer:
        if tw_df["Work Order"].count() != 0:
            tw_df.to_excel(
                writer,
                sheet_name="Scheduled for Shift",
                index=False
            )
        if mc_df["Work Order"].count() != 0:
            mc_df.to_excel(
                writer,
                sheet_name="Due by Midnight",
                index=False
            )
        if blog_df["Work Order"].count() != 0:
            blog_df.to_excel(
                writer,
                sheet_name="Backlog",
                index=False
            )
        if fu_df["Work Order"].count() != 0:
            fu_df.to_excel(
                writer,
                sheet_name="Follow Ups",
                index=False
            )
        if tr_df["Work Order"].count() != 0:
            tr_df.to_excel(
                writer,
                sheet_name="Training",
                index=False
            )


def format_work_order_summary(report_path, sheet_to_columns):
    # use openpyxl to format the excel file
    wo_summary_report = load_workbook(report_path)

    # loop though the workbook sheets
    for sheetname in wo_summary_report.sheetnames:
        sheet = wo_summary_report[sheetname]
        columns = sheet_to_columns[sheetname]
        centered_cols = [i for i in range(len(columns)) if columns[i] in [
            'Work Order', 'Sched. Start Date', 'PM Compliance Min', 'PM Compliance Max', 'Reported By']]
        left_indent_cols = [i for i in range(len(columns)) if columns[i] in [
            'Description', 'Type', 'Equipment', '1 - WO Owner', 'Assigned to on Activity (Top 8 activities)']]
        date_cols = [i for i in range(len(columns)) if columns[i] in [
            'Sched. Start Date', 'PM Compliance Min', 'PM Compliance Max']]

        # configure alignment for the sheet
        for row in sheet[2:sheet.max_row]:
            for i in centered_cols:
                cell = row[i]
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
            for i in left_indent_cols:
                cell = row[i]
                cell.alignment = Alignment(
                    horizontal='left', vertical='center', indent=1)

        for i, cell in enumerate(sheet[1]):
            if i in left_indent_cols:
                cell.alignment = Alignment(
                    horizontal='left', vertical='center', indent=1, wrap_text=True)
            else:
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)

        # set the number format of the date columns
        for row in sheet[2:sheet.max_row]:
            for i in date_cols:
                row[i].number_format = "m/d/yyyy"

        # set the header row height
        sheet.row_dimensions[1].height = 28

        # set the background color and font size of the header
        for cell in sheet[1]:
            cell.font = Font(
                bold=True,
                color='FFFFFFFF'
            )
            cell.fill = PatternFill(
                fill_type='solid',
                fgColor='FF244062'
            )

        # set the borders of the table
        thin_border = Side(border_style='thin', color='FF000000')

        for row in sheet[1:sheet.max_row]:
            try:
                if row.value:
                    row.border = Border(
                        left=thin_border,
                        right=thin_border,
                        top=thin_border,
                        bottom=thin_border
                    )
            except:
                for cell in row:
                    cell.border = Border(
                        left=thin_border,
                        right=thin_border,
                        top=thin_border,
                        bottom=thin_border
                    )

        # autofit the column width
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 5

        # disable grid lines
        sheet.sheet_view.showGridLines = False

        # copy my work for the shift to clipboard
        if sheetname == "Scheduled for Shift":
            my_work = ""
            for row in sheet[2:sheet.max_row]:
                if os.getenv("INFOR_USERNAME") in str(row[3].value):
                    for i in [0, 2]:
                        my_work += (str(row[i].value) + "\t")
                    my_work += "\n"
            pyperclip.copy(my_work)

        # freeze the first row
        sheet.freeze_panes = "A2"

        # setup a filter
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

    # save and close the file
    wo_summary_report.save(report_path)
    wo_summary_report.close()
